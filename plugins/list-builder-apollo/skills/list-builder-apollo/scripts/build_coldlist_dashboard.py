#!/usr/bin/env python3
"""Portable Bingley cold-list dashboard builder (skill-bundled).
Usage: python3 build_coldlist_dashboard.py <master.xlsx> <out.html> [<user-visible master path>]
The optional 3rd arg is the master path embedded in the page for the Download Excel
browser-download fallback; it defaults to <master.xlsx>.
Reads the FINISHED master READ-ONLY (never writes it). Prints a JSON status line incl. artifact_id.
"""
import sys, os, re, json, base64, datetime

# --- template source: editable assets, loaded at build time (kept separate per HQ standard) ---
_ASSET = os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets")
def _load(n): return open(os.path.join(_ASSET, n), encoding="utf-8").read()
try:
    SHELL  = _load("coldlist-shell.html")   # Bingley desk shell + CSS
    CENTER = _load("coldlist-center.html")  # header + boxes + table markup
    SCRIPT = _load("coldlist.js")           # data wiring, filters, sort, download
    _ASSETS_OK = True
except Exception as _e:
    SHELL = CENTER = SCRIPT = ""; _ASSETS_OK = False; _ASSET_ERR = "template assets missing: %s" % _e
EMBED_CAP=2_500_000   # bytes of raw xlsx; above this, link instead of embed
ROW_CAP=2500

def slug(s):
    s=re.sub(r"[^\w\s-]","",s or ""); s=re.sub(r"\s+","_",s.strip()); return re.sub(r"_+","_",s).strip("_")

def main():
    master, out = sys.argv[1], sys.argv[2]
    try:
        import openpyxl
    except Exception:
        print(json.dumps({"ok":False,"reason":"openpyxl missing"})); return
    if not _ASSETS_OK:
        print(json.dumps({"ok":False,"reason":_ASSET_ERR})); return
    wb=openpyxl.load_workbook(master, data_only=True)   # full load (read hidden hyperlinks); never saved
    ws=next((s for s in wb.worksheets if s.title.endswith("— Merged")),
       next((s for s in wb.worksheets if "Merged" in s.title), wb.worksheets[0]))
    hdr=[ (c.value if c.value is not None else "") for c in ws[1] ]
    idx={h:i for i,h in enumerate(hdr)}
    def col(name): return idx.get(name)
    li_c, web_c = col("LinkedIn"), col("Company")
    rows=[]
    for r in ws.iter_rows(min_row=2):
        if all((c.value is None) for c in r): continue
        d={}
        for i,h in enumerate(hdr):
            if not h: continue
            v=r[i].value
            d[h]="" if v is None else (v.strftime("%Y-%m-%d") if hasattr(v,"strftime") else str(v).strip())
        d["_li"]=""; d["_web"]=""
        def _safe_href(t): return t if isinstance(t,str) and t.lower().lstrip().startswith(("http://","https://")) else ""
        if li_c is not None and r[li_c].hyperlink and r[li_c].hyperlink.target: d["_li"]=_safe_href(r[li_c].hyperlink.target)
        if web_c is not None and r[web_c].hyperlink and r[web_c].hyperlink.target: d["_web"]=_safe_href(r[web_c].hyperlink.target)
        rows.append(d)
    # niche / geo / currency from custom properties, fallbacks
    props={}
    try:
        for p in wb.custom_doc_props.props: props[p.name]=p.value
    except Exception: pass
    niche=str(props.get("run_niche") or ws.title.replace("— Merged","").replace("Merged","").strip(" —") or "Prospects")
    geo=str(props.get("run_geography") or "").strip()
    def detect_cur():
        cg=(geo or "").lower()
        if cg in ("uk","gb","united kingdom","england","scotland","wales"): return "£"
        if cg in ("us","usa","united states","canada"): return "$"
        for d in rows:
            for sym in ("£","$","€"):
                if sym in (d.get("Rev Band") or ""): return sym
        return "£"
    currency=detect_cur()
    def live(colname, mn=1):
        if colname not in idx: return False
        vals=[ (d.get(colname) or "").strip() for d in rows ]
        nonempty=[v for v in vals if v and v.lower() not in ("unknown","")]
        return len(set(nonempty))>=mn
    show={"rev":live("Rev Band"),"size":("Employees" in idx and any((d.get("Employees") or "").strip() for d in rows)),
          "sen":("Title" in idx and len({ (d.get("Title") or "").strip() for d in rows if (d.get("Title") or "").strip()})>=1),
          "mx":live("MX Status")}
    data=open(master,"rb").read(); size=len(data); dl_name=os.path.basename(master)
    user_path=(sys.argv[3] if len(sys.argv)>3 else master)
    if size<=EMBED_CAP:
        dl_mode="embed"; xb64=base64.b64encode(data).decode()
    else:
        dl_mode="path"; xb64=""
    cfg={"niche":niche,"geo":geo,"currency":currency,"show":show,
         "dl":{"mode":dl_mode,"name":dl_name,"path":user_path},"rowCap":ROW_CAP,"total":len(rows)}
    center=CENTER.replace("__NICHE__", (niche.replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")))
    data_json=json.dumps(rows, ensure_ascii=False).replace("</","<\\/")
    script=(SCRIPT.replace("__DATA_JSON__",data_json).replace("__XLSX_B64__",xb64)
                  .replace("__CFG_JSON__",json.dumps(cfg, ensure_ascii=False).replace("</","<\\/")))
    html=(SHELL.replace("__CENTER__",center).replace("__SCRIPT__",script))
    open(out,"w",encoding="utf-8").write(html)
    aid=("cold-list-"+slug(niche)+("-"+slug(geo) if geo else "")).lower()  # lowercase: Cowork artifact ids are lowercased on store, so match the printed id to the stored id (update-in-place)
    print(json.dumps({"ok":True,"artifact_id":aid,"niche":niche,"geo":geo,"currency":currency,
        "rows":len(rows),"show":show,"dl_mode":dl_mode,"xlsx_bytes":size,"html_bytes":len(html),"out":out}))

if __name__=="__main__":
    main()
